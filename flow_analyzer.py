import pandas as pd
import os, re
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

def read_alipay(file_path):
    df = pd.read_csv(file_path, skiprows=24, encoding='gbk')  # 跳过前24行
    df = df.replace('\*', '', regex=True)
    # print(f"Alipay columns: {df.columns}")  # 打印列名以检查实际列名
    df = df.rename(columns={
        '交易时间': 'date',
        '交易分类': 'category',
        '交易对方': 'counterparty',
        '商品说明': 'description',
        '收/支': 'type',
        '金额': 'amount',
        '收/付款方式': 'payment_method',
        '交易状态': 'status'
    })
    df['source'] = 'Alipay'  # 新增source列
    df['amount'] = df['amount'].astype(float)
    df['date'] = pd.to_datetime(df['date'])
    df['category'] = '其他'  # 清除category内容
    df['detained'] = df['counterparty'] + '_' + df['description']  # 新增detained列
    return df

# 处理退款金额
def adjust_amount(row):
    if '已退款' in row['status']:
        refund_amount = float(re.search(r'￥(\d+\.\d+)', row['status']).group(1))
        return row['amount'] - refund_amount
    return row['amount']

def read_wechat(file_path):
    df = pd.read_csv(file_path, skiprows=16, encoding='utf-8')  # 跳过前16行
    df = df.replace('\*', '', regex=True)
    # print(f"WeChat columns: {df.columns}")  # 打印列名以检查实际列名
    df = df.rename(columns={
        '交易时间': 'date',
        '交易类型': 'category',
        '交易对方': 'counterparty',
        '商品': 'description',
        '收/支': 'type',
        '金额(元)': 'amount',
        '支付方式': 'payment_method',
        '当前状态': 'status'
    })
    df['source'] = 'WeChat'  # 新增source列
    df['amount'] = df['amount'].str.replace('¥', '').astype(float)
    df['amount'] = df.apply(adjust_amount, axis=1)
    df['date'] = pd.to_datetime(df['date'])
    df['category'] = '其他'  # 清除category内容
    df['detained'] = df['counterparty'] + '_' + df['description']  # 新增detained列
    return df

def parse_amount(amount_str):
    match = re.match(r'(\d+(\.\d+)?)(?:\(已退款(\d+(\.\d+)?)\))?', amount_str)
    if match:
        amount = float(match.group(1))
        refund = float(match.group(3)) if match.group(3) else 0.0
        return amount - refund
    return float(amount_str) if re.match(r'^\d+(\.\d+)?$', amount_str) else 0.0

def read_jd(file_path):
    df = pd.read_csv(file_path, skiprows=21, encoding='utf-8', sep=',', on_bad_lines='warn')  # 使用 utf-8 编码，指定分隔符为逗号，忽略异常行
    # print(f"JD columns: {df.columns}")  # 打印列名以检查实际列名
    df = df.replace('\t', '', regex=True)
    df = df.replace('\*', '', regex=True)
    df = df.rename(columns={
        '交易时间': 'date',
        '商户名称': 'counterparty',
        '交易说明': 'description',
        '金额': 'amount',
        '收/付款方式': 'payment_method',
        '交易状态': 'status',
        '收/支': 'type',
        '交易分类': 'category'
    })
    df['source'] = 'JD'  # 新增source列
    df['amount'] = df['amount'].apply(parse_amount)  # 解析金额字符串并计算实际金额
    df['date'] = df['date'].str.extract(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})')[0]  # 提取日期字符串的前面部分
    df['date'] = pd.to_datetime(df['date'], errors='coerce')  # 将无效日期转换为 NaT
    invalid_dates = df[df['date'].isna()]  # 筛选出包含无效日期的行
    if not invalid_dates.empty:
        print("Invalid dates:")
        print(invalid_dates)
    df = df.dropna(subset=['date'])  # 删除包含无效日期的行
    df['category'] = '其他'  # 清除category内容
    df['detained'] = df['counterparty'] + '_' + df['description']  # 新增detained列
    return df

def load_family_accounts(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        family_accounts = [line.strip() for line in file.readlines()]
    return family_accounts

def filter_invalid_info(df):
    if df['source'].iloc[0] == 'WeChat':
        df.loc[df['description'].str.contains('京东-订单编号', na=False), 'category'] = 'ignore'  # 将包含京东-订单编号的记录的category设置为ignore
    df.loc[df['type'].str.contains('不计收支|^/$', na=False), 'category'] = 'ignore'  # 将不计收支或/的记录的category设置为ignore
    df.loc[df['status'].str.contains('交易关闭|对方已退还|已全额退款'), 'category'] = 'ignore'  # 将交易关闭或包含退款的记录的category设置为ignore
    df.loc[df['amount'] == 0, 'category'] = 'ignore'  # 将amount为0的记录的category设置为ignore
    family_accounts = load_family_accounts('family_accounts.txt')
    df.loc[df['counterparty'].isin(family_accounts), 'category'] = 'ignore'  # 如果counterparty是family_accounts中的一个元素，则category设置为ignore
    return df

def load_categories(file_path):
    categories = {}
    with open(file_path, 'r', encoding='utf-8') as file:
        for line in file:
            category, keywords = line.strip().split(':')
            categories[category] = keywords
    return categories

def reclassify_category(df):
    categories = load_categories('categories.txt')
    conditions = [df['detained'].str.contains(keywords, na=False) for keywords in categories.values()]
    category_names = list(categories.keys())

    # 应用条件进行重新分类
    for condition, category in zip(conditions, category_names):
        df.loc[condition, 'category'] = category
    return df

def read_all_files(directory, read_function):
    all_files = [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.csv')]
    df_list = [read_function(file) for file in all_files]
    combined_df = pd.concat(df_list, ignore_index=True)
    return combined_df

def merge_data(alipay_df, wechat_df, jd_df):
    combined_df = pd.concat([alipay_df, wechat_df, jd_df], ignore_index=True)
    combined_df['month'] = combined_df['date'].dt.to_period('M')
    combined_df = reclassify_category(combined_df)
    combined_df = filter_invalid_info(combined_df)
    return combined_df

def analyze_data(df):
    df = df[df['amount'] != 0]
    monthly_expense = df[df['type'] == '支出'].groupby('month')['amount'].sum()
    category_expense = df[df['type'] == '支出'].groupby('category')['amount'].sum()
    counterparty_expense = df[df['type'] == '支出'].groupby('counterparty')['amount'].sum()
    return monthly_expense, category_expense, counterparty_expense

def save_to_excel(df, output_file, monthly_expense, category_expense):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 将 monthly_expense 和 category_expense 写入一个工作表中
        summary_df = pd.DataFrame({
            'Month': monthly_expense.index,
            'Expense': monthly_expense.values
        })
        category_df = pd.DataFrame({
            'Category': category_expense[category_expense.index != 'ignore'].index,
            'Expense': category_expense[category_expense.index != 'ignore'].values
        })
        summary_df.to_excel(writer, sheet_name='Summary', startrow=0, index=False)
        category_df.to_excel(writer, sheet_name='Summary', startrow=len(summary_df) + 2, index=False)

        ws = writer.book['Summary']
        # 创建月度支出直方图
        chart = BarChart()
        chart.title = "Monthly Expense"
        chart.x_axis.title = "Month"
        chart.y_axis.title = "Expense"

        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(summary_df) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(summary_df) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "E2")

        # 创建分类支出扇形图
        pie_chart = PieChart()
        pie_chart.title = "Category Expense"

        pie_data = Reference(ws, min_col=2, min_row=len(summary_df) + 4, max_col=2, max_row=len(summary_df) + len(category_df) + 3)
        pie_categories = Reference(ws, min_col=1, min_row=len(summary_df) + 5, max_row=len(summary_df) + len(category_df) + 3)
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_categories)
        # 设置数据标签显示百分比，保留一位小数
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showVal = False
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.numFmt = '0.0%'
        ws.add_chart(pie_chart, "E20")

        # 每月支出
        for month, group in df[(df['type'] == '支出') & (df['category'] != 'ignore')].groupby('month'):
            group.to_excel(writer, sheet_name=str(month), index=False)
            worksheet = writer.sheets[str(month)]
            worksheet.cell(row=len(group) + 2, column=group.columns.get_loc('amount') + 1, value='总计')
            worksheet.cell(row=len(group) + 2, column=group.columns.get_loc('amount') + 2, value=group['amount'].sum())

        # 不同分类的支出
        for category, group in df[(df['type'] == '支出') & (df['category'] != 'ignore')].groupby('category'):
            if not group.empty:
                sheet_name = category if category else '其他'
                group.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                worksheet.cell(row=len(group) + 2, column=group.columns.get_loc('amount') + 1, value='总计')
                worksheet.cell(row=len(group) + 2, column=group.columns.get_loc('amount') + 2, value=group['amount'].sum())

        ignore_group = df[df['category'] == 'ignore']
        if not ignore_group.empty:
            ignore_group.to_excel(writer, sheet_name='ignore', index=False)

def main():
    alipay_directory = 'alipay_files'
    wechat_directory = 'wechat_files'
    jd_directory = 'jd_files'

    alipay_df = read_all_files(alipay_directory, read_alipay)
    wechat_df = read_all_files(wechat_directory, read_wechat)
    jd_df = read_all_files(jd_directory, read_jd)

    combined_df = merge_data(alipay_df, wechat_df, jd_df)

    monthly_expense, category_expense, counterparty_expense = analyze_data(combined_df)

    print("每月开支:")
    print(monthly_expense)
    print("\n不同分类的开支:")
    print(category_expense)
    # print("\n不同人的开支情况:")
    # print(counterparty_expense)

    output_file = 'expense_analysis.xlsx'
    save_to_excel(combined_df, output_file, monthly_expense, category_expense)
    print(f"结果已保存到 {output_file}")

if __name__ == "__main__":
    main()